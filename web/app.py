"""
Minimal Flask app using domain layer.
Business logic moved to web/domain.py, routes organized in web/routes/
"""

import datetime
import io
import json
import os
import re
import sys
import uuid
from pathlib import Path

from flask import (
    Flask,
    flash,
    jsonify,
    make_response,
    redirect,
    render_template,
    request,
    url_for,
)
from pydantic import ValidationError
from werkzeug.utils import secure_filename

from .domain import (
    Configuration,
    ExcelValidator,
    FileManager,
    FiledropRouter,
)
from .models import ExcelRow, ExcelUploadRequest, JsonUploadRequest
from .utils import fill_xml_template

# ============================================================================
# FLASK APP SETUP
# ============================================================================

_MEIPASS = getattr(sys, "_MEIPASS", None)
if getattr(sys, "frozen", False) and _MEIPASS:
    _candidates = [Path(_MEIPASS) / "web", Path(_MEIPASS)]
    _WEB_ROOT = next(
        (
            p
            for p in _candidates
            if (p / "templates").exists() and (p / "static").exists()
        ),
        Path(_MEIPASS) / "web",
    )
else:
    _WEB_ROOT = Path(__file__).parent

app = Flask(
    __name__,
    template_folder=str(_WEB_ROOT / "templates"),
    static_folder=str(_WEB_ROOT / "static"),
)

# ============================================================================
# CONFIGURATION & DEPENDENCY INJECTION
# ============================================================================

CONFIG_PATH = Path(__file__).parent / "instellingen.json"
CONFIG = Configuration.load_from_file(CONFIG_PATH)

ROUTER = FiledropRouter(CONFIG)
FILE_MANAGER = FileManager(ROUTER)
RESULTS_PANEL_LIMIT = 100

ALLOWED_EXTENSIONS = {"xlsx", "xls", "xlsm"}

# Minimal app config
app.secret_key = os.environ.get("U_XMLATOR_SECRET", "xmlator-dev-secret")
app.config["MAX_CONTENT_LENGTH"] = (
    max(1, int(CONFIG.get("upload_max_size_mb", 10))) * 1024 * 1024
)

# ============================================================================
# ROUTES - PAGES
# ============================================================================


@app.route("/")
def home():
    return redirect(url_for("genereer_xml"))


@app.route("/genereer_xml")
def genereer_xml():
    files, total_count = FILE_MANAGER.list_generated_files(
        limit=RESULTS_PANEL_LIMIT, prune=False
    )
    generated = [
        {"filename": f.filename, "tijdstip": f.tijdstip, "size": f.size} for f in files
    ]
    return render_template(
        "genereer_xml.html",
        generated=generated,
        total_count=total_count,
    )


@app.route("/resultaten/fragment")
def resultaten_fragment():
    files, total_count = FILE_MANAGER.list_generated_files(
        limit=RESULTS_PANEL_LIMIT, prune=False
    )
    generated = [
        {"filename": f.filename, "tijdstip": f.tijdstip, "size": f.size} for f in files
    ]
    return make_response(
        render_template(
            "_results_panel.html",
            generated=generated,
            total_count=total_count,
        )
    )


# ============================================================================
# ROUTES - FILE MANAGEMENT
# ============================================================================


@app.route("/resultaten/download/<filename>")
def download_generated(filename):
    """Download a generated XML file"""
    if not filename.endswith(".xml") or "/" in filename or ".." in filename:
        flash("Invalid filename.", "danger")
        return redirect(request.referrer or url_for("genereer_xml"))

    directories = [
        ROUTER.get_output_directory("ZBM"),
        ROUTER.get_output_directory("Digipoort"),
        ROUTER.get_output_directory(),
    ]

    for output_dir in directories:
        if output_dir.exists():
            file_path = output_dir / filename
            if file_path.exists():
                from flask import send_file

                response = send_file(
                    file_path,
                    as_attachment=True,
                    download_name=filename,
                    mimetype="application/xml",
                )
                response.headers["X-Content-Type-Options"] = "nosniff"
                response.headers["Cache-Control"] = "no-store"
                return response

    flash("File not found.", "danger")
    return redirect(request.referrer or url_for("genereer_xml"))


@app.route("/resultaten/delete-selected", methods=["POST"])
def delete_selected_files():
    """Delete selected files"""
    try:
        data = request.get_json(silent=True) or {}
        filenames = data.get("filenames") or []

        if not isinstance(filenames, list) or not filenames:
            return jsonify({"error": "No files selected"}), 400

        deleted, missing = FILE_MANAGER.delete_files(filenames)
        return jsonify({"success": True, "deleted": deleted, "missing": missing})

    except Exception as e:
        return jsonify({"error": str(e)}), 500


@app.route("/resultaten/download-zip", methods=["POST"])
def download_generated_zip():
    """Download selected files as ZIP"""
    try:
        import zipfile

        data = request.get_json(silent=True) or {}
        filenames = data.get("filenames") or []

        if not isinstance(filenames, list) or not filenames:
            return jsonify({"error": "No files selected"}), 400

        # Build file map
        directories = [
            ROUTER.get_output_directory("ZBM"),
            ROUTER.get_output_directory("Digipoort"),
            ROUTER.get_output_directory(),
        ]
        unique_dirs = list(dict.fromkeys(directories))

        file_map = {}
        for out_dir in unique_dirs:
            if out_dir.exists():
                for fname in out_dir.iterdir():
                    if fname.suffix == ".xml" and fname.is_file():
                        file_map[fname.name] = fname

        # Create ZIP
        zip_buffer = io.BytesIO()
        with zipfile.ZipFile(zip_buffer, "w", zipfile.ZIP_DEFLATED) as zf:
            for fn in filenames:
                if fn in file_map:
                    fpath = file_map[fn]
                    if fn.startswith("digipoort_"):
                        subfolder = "UwvZwMelding_MQ_V0428"
                    else:
                        subfolder = "v0428"
                    zf.write(fpath, arcname=f"{subfolder}/{fn}")

        zip_buffer.seek(0)
        from flask import send_file

        response = send_file(
            zip_buffer,
            mimetype="application/zip",
            as_attachment=True,
            download_name="results.zip",
        )
        response.headers["Cache-Control"] = "no-store"
        return response

    except Exception as e:
        return jsonify({"error": str(e)}), 500


# ============================================================================
# ROUTES - UPLOAD ENDPOINTS
# ============================================================================


def _format_pydantic_errors(errors) -> list[str]:  # type: ignore
    """Format Pydantic validation errors to readable strings"""
    formatted = []
    for err in errors:
        loc = err.get("loc") if isinstance(err, dict) else getattr(err, "loc", None)
        if isinstance(loc, (list, tuple)):
            loc_str = ".".join(str(p) for p in loc)
        elif loc:
            loc_str = str(loc)
        else:
            loc_str = "field"
        msg = err.get("msg") if isinstance(err, dict) else getattr(err, "msg", "Invalid input")
        msg = msg or "Invalid input"
        formatted.append(f"{loc_str}: {msg}" if loc_str else msg)
    return formatted


@app.route("/genereer_xml_json/upload_json", methods=["POST"])
def upload_json():
    """Upload JSON and generate XML"""
    file = request.files.get("json_file")
    is_ajax = request.headers.get("X-Requested-With") == "XMLHttpRequest"

    if not file:
        msg = "No JSON file uploaded."
        if is_ajax:
            return jsonify({"success": False, "error": msg}), 400
        flash(msg, "danger")
        return redirect(request.referrer or url_for("genereer_xml"))

    try:
        payload = json.load(io.TextIOWrapper(file.stream, encoding="utf-8"))
    except Exception as exc:
        msg = f"Invalid JSON: {exc}"
        if is_ajax:
            return jsonify({"success": False, "error": msg}), 400
        flash(msg, "danger")
        return redirect(request.referrer or url_for("genereer_xml"))

    if not isinstance(payload, dict):
        msg = "JSON payload must be an object."
        if is_ajax:
            return jsonify({"success": False, "error": msg}), 400
        flash(msg, "danger")
        return redirect(request.referrer or url_for("genereer_xml"))

    try:
        req_model = JsonUploadRequest(
            aanvraag_type=request.form.get("aanvraag_type", "ZBM"),
            validate_request=str(request.form.get("validate", "on")).lower() != "off",
            BSN=payload.get("BSN"),
            Geboortedatum=payload.get("Geboortedatum") or payload.get("geboortedatum"),
        )
    except ValidationError as exc:
        formatted = _format_pydantic_errors(exc.errors())
        msg = f"Invalid request: {'; '.join(formatted)}"
        if is_ajax:
            return jsonify({"success": False, "error": msg, "errors": formatted}), 400
        flash(msg, "danger")
        return redirect(request.referrer or url_for("genereer_xml"))

    aanvraag_type = req_model.aanvraag_type
    cd_bericht = ExcelValidator.normalize_berichttype(aanvraag_type)

    payload.setdefault("CdBerichtType", cd_bericht)
    payload.setdefault("BronApplicatie", cd_bericht)

    # Validate required fields
    bsn_value = req_model.BSN
    geb_value = req_model.Geboortedatum
    if bsn_value is not None:
        payload["BSN"] = bsn_value
    if geb_value is not None:
        payload["Geboortedatum"] = geb_value

    if not bsn_value or str(bsn_value).strip() == "":
        msg = "Missing BSN in JSON payload."
        if is_ajax:
            return jsonify({"success": False, "error": msg}), 400
        flash(msg, "danger")
        return redirect(request.referrer or url_for("genereer_xml"))

    if not geb_value or str(geb_value).strip() == "":
        msg = "Missing Geboortedatum in JSON payload."
        if is_ajax:
            return jsonify({"success": False, "error": msg}), 400
        flash(msg, "danger")
        return redirect(request.referrer or url_for("genereer_xml"))

    unique_suffix = uuid.uuid4().hex[:8]
    ref_prefix = str(
        payload.get("referentie_prefix") or payload.get("ReferentiePrefix") or ""
    ).strip()
    try:
        tree = fill_xml_template(None, payload, unique_suffix, ref_prefix=ref_prefix)
    except Exception as exc:
        msg = f"Error building XML: {exc}"
        if is_ajax:
            return jsonify({"success": False, "error": msg}), 400
        flash(msg, "danger")
        return redirect(request.referrer or url_for("genereer_xml"))

    # Validation
    if req_model.validate_request is True:
        try:
            from lxml import etree  # type: ignore[import-not-found]

            xsd_path = CONFIG.get(
                "xsd_path", "docs/UwvZwMeldingInternBody-v0428-b01.xsd"
            )
            xsd_full = Path(__file__).parent.parent / xsd_path
            if xsd_full.exists():
                schema = etree.XMLSchema(file=str(xsd_full))  # type: ignore
                ns_body = (
                    "http://schemas.uwv.nl/UwvML/Berichten/UwvZwMeldingInternBody-v0428"
                )
                uwb = tree.getroot().find(f".//{{{ns_body}}}UwvZwMeldingInternBody")
                if uwb is None:
                    raise ValueError("UwvZwMeldingInternBody missing from XML.")
                if not schema.validate(uwb):
                    errs = "; ".join(str(e.message) for e in schema.error_log)
                    raise ValueError(f"XSD validation failed: {errs}")
        except Exception as exc:
            msg = f"Validation error: {exc}"
            if is_ajax:
                return jsonify({"success": False, "error": msg}), 400
            flash(msg, "danger")
            return redirect(request.referrer or url_for("genereer_xml"))

    # Save file
    output_dir = ROUTER.get_output_directory(aanvraag_type)
    output_dir.mkdir(parents=True, exist_ok=True)

    ts = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
    prefix = "digipoort" if cd_bericht == "OTP3" else cd_bericht.lower()
    pseudonym = uuid.uuid4().hex[:12]
    clean_ref_prefix = re.sub(r"[^A-Za-z0-9_-]", "", ref_prefix)[:20]
    identifier_part = f"{clean_ref_prefix}_" if clean_ref_prefix else ""
    filename = f"{prefix}_{identifier_part}{pseudonym}_{ts}.xml"
    file_path = output_dir / filename

    try:
        tree.write(str(file_path), encoding="utf-8", xml_declaration=True)
    except Exception as exc:
        msg = f"Error saving XML: {exc}"
        if is_ajax:
            return jsonify({"success": False, "error": msg}), 500
        flash(msg, "danger")
        return redirect(request.referrer or url_for("genereer_xml"))

    msg = f"JSON processed successfully. {aanvraag_type} XML generated."
    if is_ajax:
        return (
            jsonify({"success": True, "message": msg, "filename": filename}),
            200,
        )
    flash(msg, "success")
    return redirect(url_for("genereer_xml"))


@app.route("/upload_excel", methods=["POST"])
def upload_excel():
    """Upload Excel and generate XML"""
    file = request.files.get("excel_file")
    is_ajax = request.headers.get("X-Requested-With") == "XMLHttpRequest"

    if not file:
        msg = "Geen bestand geüpload."
        if is_ajax:
            return jsonify({"success": False, "error": msg}), 400
        flash(msg, "danger")
        return redirect(request.referrer or url_for("genereer_xml"))

    _, ext = os.path.splitext(file.filename or "")
    if ext.lower() not in [f".{e}" for e in ALLOWED_EXTENSIONS]:
        msg = "Ongeldig bestandstype; alleen .xlsx en .xls zijn toegestaan."
        if is_ajax:
            return jsonify({"success": False, "error": msg}), 400
        flash(msg, "danger")
        return redirect(request.referrer or url_for("genereer_xml"))

    try:
        req_model = ExcelUploadRequest(
            aanvraag_type=request.form.get("aanvraag_type", "ZBM"),
            validate_request=str(request.form.get("validate", "on")).lower() != "off",
        )
    except ValidationError as exc:
        formatted = _format_pydantic_errors(exc.errors())
        msg = f"Invalid request: {'; '.join(formatted)}"
        if is_ajax:
            return jsonify({"success": False, "error": msg, "errors": formatted}), 400
        flash(msg, "danger")
        return redirect(request.referrer or url_for("genereer_xml"))

    if (
        request.content_length
        and request.content_length > app.config["MAX_CONTENT_LENGTH"]
    ):
        msg = (
            f"Bestand te groot (max "
            f'{app.config["MAX_CONTENT_LENGTH"] // (1024*1024)} MB).'
        )
        if is_ajax:
            return jsonify({"success": False, "error": msg}), 400
        flash(msg, "danger")
        return redirect(request.referrer or url_for("genereer_xml"))

    uploads_dir = Path(__file__).parent.parent / "uploads"
    uploads_dir.mkdir(parents=True, exist_ok=True)
    orig_filename = file.filename if file.filename else "upload.xlsx"
    filename = secure_filename(orig_filename)
    file_path = uploads_dir / filename
    file.save(str(file_path))

    def _cell_to_str(value):
        if value is None:
            return ""
        if isinstance(value, datetime.datetime):
            return value.strftime("%Y%m%d")
        if isinstance(value, datetime.date):
            return value.strftime("%Y%m%d")
        if isinstance(value, float) and value.is_integer():
            return str(int(value))
        if isinstance(value, int):
            return str(value)
        return str(value).strip()

    if req_model.validate_request:
        try:
            import openpyxl

            wb_val = openpyxl.load_workbook(str(file_path), data_only=True)
            ws_val = wb_val.active
            if ws_val is None:
                raise ValueError("Het geüploade Excel-bestand bevat geen werkblad.")

            header_row = next(ws_val.iter_rows(min_row=1, max_row=1))
            headers = [_cell_to_str(cell.value) for cell in header_row]
            if not any(headers):
                raise ValueError("Excel-bestand bevat geen kolomkoppen.")

            errors = []
            has_full_name = "Voornaam" in headers and "Achternaam" in headers
            for idx, row in enumerate(ws_val.iter_rows(min_row=2, values_only=True), start=2):
                row_values = [_cell_to_str(v) for v in row]
                if not any(row_values):
                    continue
                row_dict = {
                    headers[i]: row_values[i]
                    for i in range(min(len(headers), len(row_values)))
                    if headers[i]
                }
                try:
                    if has_full_name:
                        ExcelRow(**row_dict)
                    else:
                        JsonUploadRequest(
                            aanvraag_type=req_model.aanvraag_type,
                            validate_request=False,
                            BSN=row_dict.get("BSN"),
                            Geboortedatum=row_dict.get("Geboortedatum"),
                        )
                except ValidationError as exc:
                    formatted = _format_pydantic_errors(exc.errors())
                    details = "; ".join(formatted)
                    errors.append(f"Row {idx}: {details}")

            if errors:
                error_msg = "Validatie mislukt: " + "; ".join(errors[:5])
                if len(errors) > 5:
                    error_msg += f" (en {len(errors) - 5} meer)"
                if is_ajax:
                    return jsonify(
                        {"success": False, "error": error_msg, "errors": errors},
                    ), 400
                flash(error_msg, "danger")
                return redirect(request.referrer or url_for("genereer_xml"))
        except Exception as exc:
            msg = f"Fout bij valideren van Excel: {exc}"
            if is_ajax:
                return jsonify({"success": False, "error": msg}), 400
            flash(msg, "danger")
            return redirect(request.referrer or url_for("genereer_xml"))

    aanvraag_type = req_model.aanvraag_type
    ref_prefix = str(request.form.get("referentie_prefix", "")).strip()
    cd_override = ExcelValidator.normalize_berichttype(aanvraag_type)

    try:
        from tools.generate_from_excel import generate_from_excel_file

        output_dir = ROUTER.get_output_directory(aanvraag_type)
        output_dir.mkdir(parents=True, exist_ok=True)
        log_path = Path(__file__).parent.parent / "build" / "logs" / "generator_excel.log"

        generated_paths = generate_from_excel_file(
            str(file_path),
            str(output_dir),
            mode="single",
            log_path=str(log_path),
            data_only=True,
            cd_bericht_override=cd_override,
            ref_prefix=ref_prefix,
        )

        generated_files = [Path(p).name for p in generated_paths]
        if not generated_files:
            raise ValueError(
                "Geen XML-bestanden gegenereerd. Controleer of alle verplichte velden aanwezig zijn (BSN, Geboortedatum)."
            )
    except Exception as exc:
        msg = f"Fout bij genereren van XML uit Excel: {exc}"
        if is_ajax:
            return jsonify({"success": False, "error": msg}), 400
        flash(msg, "danger")
        return redirect(request.referrer or url_for("genereer_xml"))

        msg = (
            f"Excel-bestand succesvol geüpload en {aanvraag_type} "
            f"XML-bestanden gegenereerd."
        )

        if is_ajax:
            response_data = {
                "success": True,
                "message": msg,
                "generated_filenames": generated_files,
                "download_links": [
                    {"filename": fname, "url": f"/resultaten/download/{fname}"}
                    for fname in generated_files
                ],
            }
            response_data["message"] += f" {len(generated_files)} bestand(en) beschikbaar voor download."
            return jsonify(response_data), 200

        flash(msg, "success")

    finally:
        try:
            if file_path.exists():
                file_path.unlink()
        except Exception:
            pass

    return redirect(url_for("genereer_xml"))


# ============================================================================
# HEALTH CHECK ENDPOINTS
# ============================================================================


@app.route("/health")
def health():
    return jsonify({"status": "healthy"}), 200


@app.route("/ready")
def ready():
    status = {"status": "ready"}
    try:
        output_dir = ROUTER.get_output_directory(quiet=True)
        output_dir.mkdir(parents=True, exist_ok=True)
    except Exception:
        status["status"] = "not_ready"

    return jsonify(status), 200 if status["status"] == "ready" else 503


@app.route("/favicon.ico")
def favicon():
    return "", 204


if __name__ == "__main__":
    app.run(debug=True, host="127.0.0.1", port=5000)
