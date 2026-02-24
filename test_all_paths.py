#!/usr/bin/env python3
"""Comprehensive path testing for XMLator"""
import io
import os
import sys
import tempfile

sys.path.insert(0, ".")
from openpyxl import Workbook

from web.app import app, get_output_directory, list_generated_files


def test_empty_excel():
    """TEST 1: Empty Excel (only header, no data rows)"""
    print("=" * 60)
    print("TEST 1: Empty Excel (only header, no data rows)")
    print("=" * 60)
    wb = Workbook()
    ws = wb.active
    ws.append(["BSN", "Voornaam", "Achternaam", "Geboortedatum"])

    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
        empty_path = f.name
        wb.save(empty_path)

    client = app.test_client()
    with open(empty_path, "rb") as f:
        resp = client.post(
            "/upload_excel",
            data={
                "excel_file": (io.BytesIO(f.read()), "empty.xlsx"),
                "aanvraag_type": "ZBM",
            },
            headers={"X-Requested-With": "XMLHttpRequest"},
        )

    print(f"Status: {resp.status_code}")
    result = resp.get_json()
    print(f'Success: {result.get("success")}')
    print(f'Message: {result.get("message")}')
    os.unlink(empty_path)
    print()


def test_single_row():
    """TEST 2: Single row Excel"""
    print("=" * 60)
    print("TEST 2: Single row Excel")
    print("=" * 60)
    wb = Workbook()
    ws = wb.active
    ws.append(["BSN", "Voornaam", "Achternaam", "Geboortedatum"])
    ws.append(["555501759", "Jan", "Jansen", "1980-12-05"])

    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
        single_path = f.name
        wb.save(single_path)

    client = app.test_client()
    with open(single_path, "rb") as f:
        resp = client.post(
            "/upload_excel",
            data={
                "excel_file": (io.BytesIO(f.read()), "single.xlsx"),
                "aanvraag_type": "ZBM",
            },
            headers={"X-Requested-With": "XMLHttpRequest"},
        )

    print(f"Status: {resp.status_code}")
    result = resp.get_json()
    print(f'Success: {result.get("success")}')
    print(f'Message: {result.get("message")}')
    os.unlink(single_path)
    print()


def test_multiple_rows():
    """TEST 3: Multiple rows Excel (3 data rows)"""
    print("=" * 60)
    print("TEST 3: Multiple rows Excel (3 data rows)")
    print("=" * 60)
    wb = Workbook()
    ws = wb.active
    ws.append(["BSN", "Voornaam", "Achternaam", "Geboortedatum"])
    ws.append(["111111111", "Alice", "Ander", "1975-01-01"])
    ws.append(["222222222", "Bob", "Baker", "1980-02-02"])
    ws.append(["333333333", "Charlie", "Chen", "1985-03-03"])

    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
        multi_path = f.name
        wb.save(multi_path)

    client = app.test_client()
    with open(multi_path, "rb") as f:
        resp = client.post(
            "/upload_excel",
            data={
                "excel_file": (io.BytesIO(f.read()), "multi.xlsx"),
                "aanvraag_type": "ZBM",
            },
            headers={"X-Requested-With": "XMLHttpRequest"},
        )

    print(f"Status: {resp.status_code}")
    result = resp.get_json()
    print(f'Success: {result.get("success")}')
    print(f'Message: {result.get("message")}')
    os.unlink(multi_path)
    print()


def test_digipoort():
    """TEST 4: Digipoort aanvraag type"""
    print("=" * 60)
    print("TEST 4: Digipoort aanvraag type")
    print("=" * 60)
    wb = Workbook()
    ws = wb.active
    ws.append(["BSN", "Voornaam", "Achternaam", "Geboortedatum"])
    ws.append(["444444444", "Diana", "Davis", "1990-04-04"])

    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
        digi_path = f.name
        wb.save(digi_path)

    client = app.test_client()
    with open(digi_path, "rb") as f:
        resp = client.post(
            "/upload_excel",
            data={
                "excel_file": (io.BytesIO(f.read()), "digi.xlsx"),
                "aanvraag_type": "Digipoort",
            },
            headers={"X-Requested-With": "XMLHttpRequest"},
        )

    print(f"Status: {resp.status_code}")
    result = resp.get_json()
    print(f'Success: {result.get("success")}')
    print(f'Message: {result.get("message")}')
    os.unlink(digi_path)
    print()


def test_list_files():
    """TEST 5: List generated files"""
    print("=" * 60)
    print("TEST 5: List generated files (no pruning)")
    print("=" * 60)
    generated, total = list_generated_files(limit=25, prune=False)
    print(f"Total files counted: {total}")
    print(f"Files returned: {len(generated)}")
    if generated:
        print("Sample files:")
        for f in generated[:5]:
            print(f'  - {f["filename"]} ({f["size"]} bytes, {f["tijdstip"]})')
    print()


def test_directories():
    """TEST 6: Output directory resolution"""
    print("=" * 60)
    print("TEST 6: Output directory resolution")
    print("=" * 60)
    for atype in ["ZBM", "Digipoort", None]:
        path = get_output_directory(atype)
        exists = os.path.exists(path)
        num_files = (
            len([f for f in os.listdir(path) if f.endswith(".xml")]) if exists else 0
        )
        label = atype or "Default"
        print(f"{label:15} -> {path}")
        print(f'{"":15}    {num_files} XML files, exists={exists}')
    print()


def test_delete_selected():
    """TEST 7: Delete selected files"""
    print("=" * 60)
    print("TEST 7: Delete selected files API")
    print("=" * 60)
    generated, _ = list_generated_files(limit=5, prune=False)
    if not generated:
        print("  No files to test deletion")
        print()
        return

    client = app.test_client()
    filenames = [generated[0]["filename"]]
    print(f"  Deleting: {filenames[0]}")

    resp = client.post(
        "/resultaten/delete-selected",
        json={"filenames": filenames},
        headers={"Content-Type": "application/json"},
    )

    print(f"  Status: {resp.status_code}")
    result = resp.get_json()
    print(f'  Success: {result.get("success")}')
    print(f'  Deleted: {result.get("deleted")}')
    print()


def test_download_zip():
    """TEST 8: Download ZIP functionality"""
    print("=" * 60)
    print("TEST 8: Download ZIP functionality")
    print("=" * 60)
    generated, _ = list_generated_files(limit=3, prune=False)
    if not generated:
        print("  No files to test ZIP download")
        print()
        return

    client = app.test_client()
    filenames = [f["filename"] for f in generated[:2]]
    print(f"  Zipping {len(filenames)} files")

    resp = client.post(
        "/resultaten/download-zip",
        json={"filenames": filenames},
        headers={"Content-Type": "application/json"},
    )

    print(f"  Status: {resp.status_code}")
    print(f"  Content-Type: {resp.content_type}")
    print(f"  ZIP size: {len(resp.data)} bytes")
    if resp.status_code == 200:
        print("  [OK] ZIP download works")
    print()


if __name__ == "__main__":
    test_empty_excel()
    test_single_row()
    test_multiple_rows()
    test_digipoort()
    test_list_files()
    test_directories()
    test_delete_selected()
    test_download_zip()

    print("=" * 60)
    print("[OK] All comprehensive path tests completed")
    print("=" * 60)
