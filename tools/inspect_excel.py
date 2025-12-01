from pathlib import Path
import sys
from openpyxl import load_workbook

EXCEL = Path(__file__).parent.parent / 'docs' / 'Input XML electr ziekmeldinge.xlsx'
if not EXCEL.exists():
    print('Excel not found:', EXCEL)
    sys.exit(1)

wb = load_workbook(EXCEL, read_only=True, data_only=True)
s = wb.active
it = s.iter_rows(values_only=True)
try:
    headers = next(it)
except StopIteration:
    print('Empty sheet')
    sys.exit(1)

print('Headers:')
for i,h in enumerate(headers, start=1):
    print(i, repr(h))

print('\nFirst 6 data rows:')
for r, row in zip(range(2, 8), it):
    print('Row', r, row)

print('\nDone')
