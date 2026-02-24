#!/usr/bin/env python3
"""Adversarial/chaos testing - try to break the system intentionally"""
import io
import os
import sys
import tempfile

sys.path.insert(0, ".")
from openpyxl import Workbook

from web.app import app, list_generated_files


def test_missing_bsn():
    """TEST 1: Excel without BSN column"""
    print("=" * 60)
    print("TEST 1: Excel without BSN column (required field)")
    print("=" * 60)
    wb = Workbook()
    ws = wb.active
    ws.append(["Voornaam", "Achternaam", "Geboortedatum"])  # NO BSN!
    ws.append(["Jan", "Jansen", "1980-12-05"])

    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
        path = f.name
        wb.save(path)

    client = app.test_client()
    with open(path, "rb") as f:
        resp = client.post(
            "/upload_excel",
            data={
                "excel_file": (io.BytesIO(f.read()), "nobsn.xlsx"),
                "aanvraag_type": "ZBM",
            },
            headers={"X-Requested-With": "XMLHttpRequest"},
        )

    print(f"Status: {resp.status_code}")
    result = resp.get_json()
    print(f'Success: {result.get("success")}')
    if not result.get("success"):
        print(f'ERROR (expected): {result.get("error", result.get("message"))}')
    else:
        print("[BUG] Should have failed but didn't!")
    os.unlink(path)
    print()


def test_missing_geboortedatum():
    """TEST 2: Excel without Geboortedatum (required)"""
    print("=" * 60)
    print("TEST 2: Excel without Geboortedatum (required field)")
    print("=" * 60)
    wb = Workbook()
    ws = wb.active
    ws.append(["BSN", "Voornaam", "Achternaam"])  # NO Geboortedatum!
    ws.append(["555501759", "Jan", "Jansen"])

    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
        path = f.name
        wb.save(path)

    client = app.test_client()
    with open(path, "rb") as f:
        resp = client.post(
            "/upload_excel",
            data={
                "excel_file": (io.BytesIO(f.read()), "nogeb.xlsx"),
                "aanvraag_type": "ZBM",
            },
            headers={"X-Requested-With": "XMLHttpRequest"},
        )

    print(f"Status: {resp.status_code}")
    result = resp.get_json()
    print(f'Success: {result.get("success")}')
    if not result.get("success"):
        print(f'ERROR (expected): {result.get("error", result.get("message"))}')
    else:
        print("[BUG] Should have failed but didn't!")
    os.unlink(path)
    print()


def test_invalid_bsn():
    """TEST 3: Excel with invalid BSN (not 8-9 digits)"""
    print("=" * 60)
    print("TEST 3: Invalid BSN format (letters instead of digits)")
    print("=" * 60)
    wb = Workbook()
    ws = wb.active
    ws.append(["BSN", "Voornaam", "Achternaam", "Geboortedatum"])
    ws.append(["ABCDEFGH", "Jan", "Jansen", "1980-12-05"])  # Invalid BSN

    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
        path = f.name
        wb.save(path)

    client = app.test_client()
    with open(path, "rb") as f:
        resp = client.post(
            "/upload_excel",
            data={
                "excel_file": (io.BytesIO(f.read()), "badsn.xlsx"),
                "aanvraag_type": "ZBM",
            },
            headers={"X-Requested-With": "XMLHttpRequest"},
        )

    print(f"Status: {resp.status_code}")
    result = resp.get_json()
    print(f'Success: {result.get("success")}')
    print("[INFO] If generation succeeded, validation may not be strict enough")
    os.unlink(path)
    print()


def test_invalid_date():
    """TEST 4: Excel with invalid date format"""
    print("=" * 60)
    print("TEST 4: Invalid date format (invalid day/month)")
    print("=" * 60)
    wb = Workbook()
    ws = wb.active
    ws.append(["BSN", "Voornaam", "Achternaam", "Geboortedatum"])
    ws.append(["555501759", "Jan", "Jansen", "1980-13-32"])  # Invalid date!

    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
        path = f.name
        wb.save(path)

    client = app.test_client()
    with open(path, "rb") as f:
        resp = client.post(
            "/upload_excel",
            data={
                "excel_file": (io.BytesIO(f.read()), "baddate.xlsx"),
                "aanvraag_type": "ZBM",
            },
            headers={"X-Requested-With": "XMLHttpRequest"},
        )

    print(f"Status: {resp.status_code}")
    result = resp.get_json()
    print(f'Success: {result.get("success")}')
    if not result.get("success"):
        print(f'ERROR (expected): {result.get("error", result.get("message"))}')
    os.unlink(path)
    print()


def test_wrong_file_type():
    """TEST 5: Upload non-Excel file (e.g., TXT)"""
    print("=" * 60)
    print("TEST 5: Upload wrong file type (TXT instead of XLSX)")
    print("=" * 60)

    client = app.test_client()
    txt_data = b"This is not an Excel file"
    resp = client.post(
        "/upload_excel",
        data={
            "excel_file": (io.BytesIO(txt_data), "notexcel.txt"),
            "aanvraag_type": "ZBM",
        },
        headers={"X-Requested-With": "XMLHttpRequest"},
    )

    print(f"Status: {resp.status_code}")
    result = resp.get_json()
    print(f'Success: {result.get("success")}')
    if not result.get("success"):
        print(f'ERROR (expected): {result.get("error", result.get("message"))}')
    else:
        print("[BUG] Should have rejected non-Excel file!")
    print()


def test_path_traversal():
    """TEST 6: Try path traversal in delete endpoint"""
    print("=" * 60)
    print("TEST 6: Try path traversal attack in delete")
    print("=" * 60)

    client = app.test_client()
    resp = client.post(
        "/resultaten/delete-selected",
        json={"filenames": ["../../../etc/passwd.xml"]},  # Try to escape!
        headers={"Content-Type": "application/json"},
    )

    print(f"Status: {resp.status_code}")
    result = resp.get_json()
    print(f'Success: {result.get("success")}')
    print(f"Response: {result}")
    print("[INFO] Path traversal should be blocked by validation")
    print()


def test_path_traversal_download():
    """TEST 7: Try path traversal in download endpoint"""
    print("=" * 60)
    print("TEST 7: Try path traversal attack in download")
    print("=" * 60)

    client = app.test_client()
    resp = client.get("/resultaten/download/../../../../etc/passwd")

    print(f"Status: {resp.status_code}")
    if resp.status_code == 404:
        print("[OK] Path traversal blocked (404)")
    elif resp.status_code == 302 or resp.status_code == 400:
        print("[OK] Redirected or rejected")
    else:
        print(f"[WARNING] Unexpected status: {resp.status_code}")
    print()


def test_huge_file():
    """TEST 8: Very large Excel (10000 rows)"""
    print("=" * 60)
    print("TEST 8: Very large Excel (10000 rows)")
    print("=" * 60)
    wb = Workbook()
    ws = wb.active
    ws.append(["BSN", "Voornaam", "Achternaam", "Geboortedatum"])

    # Add 10000 rows
    for i in range(10000):
        bsn = 100000000 + i
        ws.append([str(bsn), f"Jan{i}", "Jansen", "1980-12-05"])

    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
        path = f.name
        wb.save(path)

    file_size_mb = os.path.getsize(path) / (1024 * 1024)
    print(f"[INFO] File size: {file_size_mb:.2f} MB")

    if file_size_mb > 16:
        print("[OK] File exceeds 16 MB upload limit, should be rejected")
        print("     (not uploading to avoid memory issues)")
    else:
        print("[INFO] Attempting upload of large file...")
        client = app.test_client()
        try:
            with open(path, "rb") as f:
                resp = client.post(
                    "/upload_excel",
                    data={
                        "excel_file": (io.BytesIO(f.read()), "huge.xlsx"),
                        "aanvraag_type": "ZBM",
                    },
                    headers={"X-Requested-With": "XMLHttpRequest"},
                )

            print(f"Status: {resp.status_code}")
            result = resp.get_json()
            print(f'Success: {result.get("success")}')
        except Exception as e:
            print(f"[INFO] Exception (expected for large file): {type(e).__name__}")

    os.unlink(path)
    print()


def test_special_characters():
    """TEST 9: Excel with special/unicode characters"""
    print("=" * 60)
    print("TEST 9: Special/unicode characters in data")
    print("=" * 60)
    wb = Workbook()
    ws = wb.active
    ws.append(["BSN", "Voornaam", "Achternaam", "Geboortedatum"])
    ws.append(["555501759", 'Jan<>&"', 'O\'Reilly & Co. "Test"', "1980-12-05"])

    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
        path = f.name
        wb.save(path)

    client = app.test_client()
    with open(path, "rb") as f:
        resp = client.post(
            "/upload_excel",
            data={
                "excel_file": (io.BytesIO(f.read()), "special.xlsx"),
                "aanvraag_type": "ZBM",
            },
            headers={"X-Requested-With": "XMLHttpRequest"},
        )

    print(f"Status: {resp.status_code}")
    result = resp.get_json()
    print(f'Success: {result.get("success")}')
    if result.get("success"):
        print("[OK] Special characters handled correctly")
    os.unlink(path)
    print()


def test_concurrent_delete():
    """TEST 10: Try deleting file twice concurrently"""
    print("=" * 60)
    print("TEST 10: Delete same file twice (race condition)")
    print("=" * 60)

    generated, _ = list_generated_files(limit=3, prune=False)
    if not generated:
        print("[INFO] No files to test with")
        print()
        return

    filename = generated[0]["filename"]
    client = app.test_client()

    print(f'[INFO] Attempting to delete "{filename}" twice...')

    # First delete
    resp1 = client.post(
        "/resultaten/delete-selected",
        json={"filenames": [filename]},
        headers={"Content-Type": "application/json"},
    )

    result1 = resp1.get_json()
    print(
        f'First delete - Status: {resp1.status_code}, Success: {result1.get("success")}'
    )

    # Second delete (file already gone)
    resp2 = client.post(
        "/resultaten/delete-selected",
        json={"filenames": [filename]},
        headers={"Content-Type": "application/json"},
    )

    result2 = resp2.get_json()
    print(
        f'Second delete - Status: {resp2.status_code}, Success: {result2.get("success")}'
    )
    print(f"Second delete response: {result2}")

    if result2.get("missing"):
        print("[OK] Correctly reported file as missing on second delete")
    print()


def test_invalid_json():
    """TEST 11: Send malformed JSON to delete endpoint"""
    print("=" * 60)
    print("TEST 11: Malformed JSON to delete endpoint")
    print("=" * 60)

    client = app.test_client()
    resp = client.post(
        "/resultaten/delete-selected",
        data="{invalid json}",  # Malformed
        headers={"Content-Type": "application/json"},
    )

    print(f"Status: {resp.status_code}")
    result = resp.get_json() if resp.is_json else resp.data
    print(f"Response: {result}")
    if resp.status_code >= 400:
        print("[OK] Malformed JSON rejected")
    print()


def test_upload_json_invalid():
    """TEST 12: Upload invalid JSON (not JSON format)"""
    print("=" * 60)
    print("TEST 12: Upload non-JSON file as JSON")
    print("=" * 60)

    client = app.test_client()
    resp = client.post(
        "/genereer_xml_json/upload_json",
        data={
            "json_file": (io.BytesIO(b"not json at all"), "notjson.txt"),
            "aanvraag_type": "ZBM",
        },
        headers={"X-Requested-With": "XMLHttpRequest"},
    )

    print(f"Status: {resp.status_code}")
    result = resp.get_json()
    print(f'Success: {result.get("success")}')
    if not result.get("success"):
        print(f'ERROR (expected): {result.get("error", result.get("message"))}')
    else:
        print("[BUG] Should have rejected invalid JSON!")
    print()


def test_empty_file_delete():
    """TEST 13: Try deleting empty filename"""
    print("=" * 60)
    print("TEST 13: Try deleting empty filename")
    print("=" * 60)

    client = app.test_client()
    resp = client.post(
        "/resultaten/delete-selected",
        json={"filenames": ["", None, "valid.xml"]},
        headers={"Content-Type": "application/json"},
    )

    print(f"Status: {resp.status_code}")
    result = resp.get_json()
    print(f'Success: {result.get("success")}')
    print(f"Response: {result}")
    print("[INFO] Invalid filenames should be skipped")
    print()


if __name__ == "__main__":
    print("[CHAOS TEST SUITE - TRYING TO BREAK THE SYSTEM]")
    print()

    test_missing_bsn()
    test_missing_geboortedatum()
    test_invalid_bsn()
    test_invalid_date()
    test_wrong_file_type()
    test_path_traversal()
    test_path_traversal_download()
    test_huge_file()
    test_special_characters()
    test_concurrent_delete()
    test_invalid_json()
    test_upload_json_invalid()
    test_empty_file_delete()

    print("=" * 60)
    print("[OK] Chaos test suite completed")
    print("=" * 60)
